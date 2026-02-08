# Standard library
import calendar
import os
import random
import time
import logging

from datetime import date, timedelta
from io import BytesIO

# Third-party (reportes)
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from xhtml2pdf import pisa

# Django
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.views import LoginView, PasswordChangeView, PasswordResetConfirmView
from django.core.cache import cache
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.core.mail import send_mail
from django.db import transaction
from django.db.models.functions import Lower
from django.db.models import Q, Sum, Count, Exists, OuterRef, Subquery 
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import get_template
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.html import format_html
from django.utils.http import url_has_allowed_host_and_scheme
from taxis.utils import registrar_movimiento, texto_filtros, log_pago, log_password_change
from django.views.decorators.csrf import csrf_protect, ensure_csrf_cookie
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, View, TemplateView
from django.contrib.auth.hashers import make_password
from django.views.decorators.http import require_http_methods

from django.contrib.staticfiles import finders
from urllib.parse import urlparse
from django.conf import settings
from django.contrib.staticfiles import finders




# Local
from .forms import (
    ConductorForm, VehiculoForm, PresidenteRegisterForm,
    EmailOrUsernameAuthenticationForm, PagoForm,
    UbicacionGeograficaForm, PagoMensualForm, PresidentePerfilForm, CooperativaLegalDocsForm, PresidenteIdentificacionDocsForm
)
from .models import (
    Conductor, Vehiculo, CustomUser, UbicacionGeografica,
    Deuda, Pago, MovimientoAudit, ConfiguracionCooperativa, ConfiguracionGlobal,
    DocumentoLegal, EmailVerificationCode,
    ConfiguracionFinanzas,
    PagoMensual, PendingPresidentRegistration, CooperativaLegalDocs, PresidenteIdentificacionDocs
)
from .utils import render_to_pdf

# --- DATOS FIJOS DE LA COOPERATIVA ---
DATOS_COOP = {
    'nombre': 'WILSON TORRES 33, R.L.',
    'rif': 'J-40126249-0',
    'direccion': 'CALLE LA GLORIA, CASA N° 115-C, SECTOR JESÚS BANDRES. SAN JUAN DE LOS MORROS - ESTADO GUÁRICO',
    'presidente': 'WILSON TORRES',
    'telefono': '0416-6444886',
    'email': 'wilsontorres27@gmail.com',
    'municipio': 'JUAN GERMAN ROSCIO',
}

def obtener_rutas_logos():
    """
    Obtiene rutas de logos absolutas para xhtml2pdf en Windows.
    Prioriza la ruta estática confirmada por el usuario.
    """
    # Ruta base del proyecto
    base_dir = settings.BASE_DIR
    
    # Directorios donde buscar
    search_dirs = [
        os.path.join(base_dir, 'static', 'img'),
        os.path.join(settings.STATIC_ROOT, 'img') if settings.STATIC_ROOT else None,
        os.path.join(base_dir, 'media', 'vehiculos'),
    ]
    
    logo_transporte = ""
    logo_mision = ""
    
    for directory in search_dirs:
        if not directory or not os.path.exists(directory):
            continue
            
        t_path = os.path.join(directory, 'logo_transporte.png')
        m_path = os.path.join(directory, 'logo_gran_mision.png')
        
        if not logo_transporte and os.path.exists(t_path):
            # En Windows xhtml2pdf a veces prefiere la ruta absoluta con / y sin file:///
            logo_transporte = t_path.replace(os.sep, '/')
            
        if not logo_mision and os.path.exists(m_path):
            logo_mision = m_path.replace(os.sep, '/')
            
        if logo_transporte and logo_mision:
            break
            
    return logo_transporte, logo_mision

# ====================================================================
# CONFIGURACIÓN Y SISTEMA (AJAX)
# ====================================================================

@login_required
@require_http_methods(["GET"])
def ajax_check_duplicado(request):
    """
    Vista AJAX para validar unicidad de:
    - cedula (cédula de identidad)
    - rif (RIF del conductor)
    - email (correo electrónico)
    - telefono (teléfono principal)
    
    Parámetros GET esperados:
    - campo: tipo de validación (cedula, rif, email, telefono)
    - valor: valor a validar
    - excludeid: ID del conductor actual (para edición, excluir de búsqueda)
    
    Retorna:
    - {'existe': True/False}
    """
    # ✅ CORRECCIÓN 1: Nombre de parámetro 'excludeid' (sin guion, minúscula)
    campo = request.GET.get('campo')
    valor = request.GET.get('valor', '').strip()
    exclude_id = request.GET.get('excludeid')  # ✅ CORREGIDO

    # Validar entrada básica
    if not campo or not valor:
        return JsonResponse({'existe': False})

    existe = False

    # Base de querysets
    qs_cond = Conductor.objects.all()
    
    # ✅ CORRECCIÓN 2: Convertir a int y manejar excepciones
    if exclude_id:
        try:
            exclude_id = int(exclude_id)
            qs_cond = qs_cond.exclude(pk=exclude_id)
        except (ValueError, TypeError):
            pass  # Si no es un número válido, ignorar la exclusión

    qs_users = CustomUser.objects.all()

    # ✅ VALIDACIONES POR TIPO DE CAMPO
    
    if campo == 'cedula':
        # ✅ CORRECCIÓN 3: Limpiar números igual que en el formulario
        val_limpio = ''.join(filter(str.isdigit, valor))
        existe = qs_cond.filter(cedula_identidad=val_limpio).exists()

    elif campo == 'rif':
        # ✅ CORRECCIÓN 4: Limpiar números para RIF
        val_limpio = ''.join(filter(str.isdigit, valor))
        existe = qs_cond.filter(rif=val_limpio).exists()

    elif campo == 'email':
        # Email se valida sin cambios (case-insensitive)
        existe_cond = qs_cond.filter(email__iexact=valor).exists()
        existe_user = qs_users.filter(email__iexact=valor).exists()
        existe = existe_cond or existe_user

    elif campo == 'telefono':
        # Limpiar teléfono y comparar últimos 10 dígitos
        val_limpio = ''.join(filter(str.isdigit, valor))
        val_comparar = val_limpio[-10:] if len(val_limpio) >= 10 else val_limpio

        # A. Revisar en Conductores
        for c in qs_cond:
            if c.telefono_principal:
                db_limpio = ''.join(filter(str.isdigit, c.telefono_principal))
                db_comparar = db_limpio[-10:] if len(db_limpio) >= 10 else db_limpio
                if db_comparar == val_comparar:
                    existe = True
                    break

        # B. Revisar en Usuarios (solo si no encontró en Conductores)
        if not existe:
            for u in qs_users:
                tlf_user = getattr(u, 'phone_number', '')
                if tlf_user:
                    db_limpio = ''.join(filter(str.isdigit, str(tlf_user)))
                    db_comparar = db_limpio[-10:] if len(db_limpio) >= 10 else db_limpio
                    if db_comparar == val_comparar:
                        existe = True
                        break

    return JsonResponse({'existe': existe})

@login_required
def buscar_chofer(request):
    """
    AJAX: Buscar chofer (Conductor) por cédula.
    Retorna:
    - encontrado: bool
    - id: int
    - nombre: str
    """
    cedula = request.GET.get('cedula', '').strip()
    if not cedula:
        return JsonResponse({'encontrado': False})

    cedula_limpia = ''.join(filter(str.isdigit, cedula))
    if not cedula_limpia:
        return JsonResponse({'encontrado': False})

    conductor = Conductor.objects.filter(cedula_identidad=cedula_limpia).first()
    if not conductor:
        return JsonResponse({'encontrado': False})

    return JsonResponse({
        'encontrado': True,
        'id': conductor.id,
        'nombre': f"{conductor.nombres} {conductor.apellidos}",
    })

@login_required
def validar_datos_vehiculo(request):
    """
    API AJAX para validar Placa, Serial y Casco en tiempo real.
    Devuelve si existe y un mensaje personalizado.
    """
    campo = request.GET.get('campo')
    valor = request.GET.get('valor', '').strip().upper()
    exclude_id = request.GET.get('exclude_id')

    if not campo or not valor:
        return JsonResponse({'existe': False})

    existe = False
    mensaje = ""
    qs = Vehiculo.objects.all()
    
    if exclude_id and exclude_id not in ['None', '']:
        qs = qs.exclude(id=exclude_id)

    if campo == 'placa':
        existe = qs.filter(placa=valor).exists()
        if existe:
            mensaje = "Esta placa ya está registrada en el sistema."
            
    elif campo == 'serial_niv':
        existe = qs.filter(serial_niv=valor).exists()
        if existe:
            mensaje = "Este serial de carrocería (NIV) ya existe."
            
    elif campo == 'numero_casco':
        existe = qs.filter(numero_casco=valor).exists()
        if existe:
            mensaje = "Este número de casco ya está asignado a otro vehículo."

    return JsonResponse({'existe': existe, 'mensaje': mensaje})


@login_required
def ejecutar_cierre_mensual(request):
    if request.user.role.upper() != 'PRESIDENTE':
        messages.error(request, "No tienes permisos para ejecutar esta acción.")
        return redirect('taxis:panel_general')

    tasa_actual = ConfiguracionGlobal.get_tasa()
    monto_bs = float(tasa_actual) * 5.00

    afiliados_activos = Conductor.objects.filter(estado='activo')
    hoy = timezone.now().date()
    count_deudas = 0
    
    with transaction.atomic():
        for afiliado in afiliados_activos:
            existe = Deuda.objects.filter(conductor=afiliado, mes=hoy.month, anio=hoy.year).exists()
            if not existe:
                Deuda.objects.create(
                    conductor=afiliado, mes=hoy.month, anio=hoy.year,
                    monto_bs=monto_bs, concepto=f"Cuota {hoy.month}/{hoy.year}",
                    pagada=False, fecha_vencimiento=hoy + timezone.timedelta(days=5)
                )
                count_deudas += 1

        fecha_limite = timezone.now() - timezone.timedelta(days=365)
        pagos_viejos = Pago.objects.filter(fecha_pago__lt=fecha_limite)
        count_borrados = pagos_viejos.count()
        pagos_viejos.delete()

    registrar_movimiento(
       request=request,
       accion=f"Ejecutó cierre mensual {count_deudas} deudas generadas",
       modulo="finanzas",
    )


    messages.success(request, f"Cierre completado: {count_deudas} deudas nuevas, {count_borrados} pagos archivados.")
    return redirect('taxis:panel_general')

# ====================================================================
# CONDUCTORES Y VEHÍCULOS
# ====================================================================

@method_decorator(never_cache, name="dispatch")
class ConductorListView(LoginRequiredMixin, ListView):
    model = Conductor
    template_name = "taxisconductorlist.html"
    context_object_name = "conductores"
    paginate_by = 10

    def get_queryset(self):
        qs = (
            Conductor.objects
            .select_related("ubicacion")
            .prefetch_related("vehiculos")
            .order_by("id")
        )

        vehiculo_operativo = (
            Vehiculo.objects
            .filter(conductor=OuterRef("pk"), condicion__isnull=False)
            .annotate(c=Lower("condicion"))
            .filter(c="operativo")
        )

        vehiculo_inoperativo = (
            Vehiculo.objects
            .filter(conductor=OuterRef("pk"), condicion__isnull=False)
            .annotate(c=Lower("condicion"))
            .filter(c="inoperativo")
        )

        qs = qs.annotate(
            tienevehiculooperativo=Exists(vehiculo_operativo),
            tienevehiculoinoperativo=Exists(vehiculo_inoperativo),
        )

        # --- BÚSQUEDA: cédula, nombres, apellidos ---
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(nombres__icontains=q) |
                Q(apellidos__icontains=q) |
                Q(cedula_identidad__icontains=q)
            )

        # --- FILTRO: género ---
        genero = (self.request.GET.get("genero") or "").strip()
        if genero in ("M", "F"):
            qs = qs.filter(sexo=genero)

        # --- FILTRO: estatus vehículo ---
        estatus = (self.request.GET.get("estatus") or "").strip().lower()
        if estatus == "operativo":
            qs = qs.filter(tienevehiculooperativo=True)
        elif estatus == "inoperativo":
            qs = qs.filter(tienevehiculoinoperativo=True)
        elif estatus in ("sinvehiculo", "sin_vehiculo"):
            qs = qs.filter(vehiculos__isnull=True)

        return qs


class ConductorCreateView(LoginRequiredMixin, CreateView):
    model = Conductor
    form_class = ConductorForm
    template_name = "taxis/conductor_form.html"
    success_url = reverse_lazy("taxis:conductor_list")  # ✅ Irá aquí después de guardar

    def get_form(self, form_class=None):
        if form_class is None:
           form_class = self.get_form_class()
        print("FORM_CLASS_REAL =", form_class, "FROM =", getattr(form_class, "__module__", None))
        print("FORM_KWARGS =", self.get_form_kwargs())
        return form_class(**self.get_form_kwargs())

 

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["is_create"] = True
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["ubicacion_form"] = UbicacionGeograficaForm(self.request.POST)
        else:
            context["ubicacion_form"] = UbicacionGeograficaForm()
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        ubicacion_form = context["ubicacion_form"]

        if form.is_valid() and ubicacion_form.is_valid():
            with transaction.atomic():
                ubicacion = ubicacion_form.save()
                self.object = form.save(commit=False)
                self.object.ubicacion = ubicacion
                self.object.save()

                registrar_movimiento(
                    request=self.request,
                    accion=f"Creó afiliado {self.object.nombres}",
                    modulo="afiliados",
                )

            messages.success(self.request, "Afiliado registrado exitosamente.")
            return super().form_valid(form)

        return self.render_to_response(self.get_context_data(form=form))


class ConductorUpdateView(LoginRequiredMixin, UpdateView):
    model = Conductor
    form_class = ConductorForm
    template_name = "taxis/conductor_form.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["is_create"] = False
        return kwargs

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        if next_url:
            return next_url
        return reverse_lazy("taxis:dt5_detail")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["ubicacion_form"] = UbicacionGeograficaForm(
                self.request.POST, instance=self.object.ubicacion
            )
        else:
            context["ubicacion_form"] = UbicacionGeograficaForm(instance=self.object.ubicacion)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        ubicacion_form = context["ubicacion_form"]

        if form.is_valid() and ubicacion_form.is_valid():
            with transaction.atomic():
                ubicacion_form.save()
                self.object = form.save()

                registrar_movimiento(
                    request=self.request,
                    accion=f"Actualizó afiliado {self.object.nombres}",
                    modulo="afiliados",
                )

            messages.success(self.request, "Afiliado actualizado.")
            return redirect(self.get_success_url())

        return self.render_to_response(self.get_context_data(form=form))


class ConductorDeleteView(LoginRequiredMixin, DeleteView):
    model = Conductor
    template_name = "taxis/conductor_confirm_delete.html"
    success_url = reverse_lazy("taxis:conductor_list")

@method_decorator(never_cache, name="dispatch")
class VehiculoListView(LoginRequiredMixin, ListView):
    model = Vehiculo
    template_name = "taxis/vehiculo_list.html"
    context_object_name = "vehiculos"
    paginate_by = 10

    def get_queryset(self):
        qs = Vehiculo.objects.select_related('conductor').order_by('id')

        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(
                Q(numero_casco__icontains=q) |
                Q(placa__icontains=q) |
                Q(marca__icontains=q) |
                Q(modelo__icontains=q) |
                Q(conductor__nombres__icontains=q) |
                Q(conductor__apellidos__icontains=q) |
                Q(conductor__cedula_identidad__icontains=q)
            )

        estado = self.request.GET.get('estado')
        if estado in ['operativo', 'inoperativo']:
            qs = qs.filter(condicion=estado)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['estado_actual'] = self.request.GET.get('estado', 'todos')
        return context


class VehiculoCreateView(LoginRequiredMixin, CreateView):
    model = Vehiculo
    form_class = VehiculoForm
    template_name = "taxis/vehiculo_form.html"
    success_url = reverse_lazy("taxis:vehiculo_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['is_create'] = True
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        conductor_id = self.request.GET.get('conductor')
        if conductor_id:
            initial['conductor'] = conductor_id
        return initial

    def form_valid(self, form):
        self.object = form.save()
        
        registrar_movimiento(
          request=self.request,
          accion=f"Vehículo creado {self.object.placa}",
          modulo="vehiculos",
        )

        messages.success(self.request, "Vehículo registrado exitosamente.")
        
        return HttpResponseRedirect(self.success_url)


class VehiculoUpdateView(LoginRequiredMixin, UpdateView):
    model = Vehiculo
    form_class = VehiculoForm
    template_name = "taxis/vehiculo_form.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['is_create'] = False
        return kwargs

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={self.request.get_host()}):
            return next_url
        return reverse("taxis:vehiculo_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        self.object = form.save()
        
        registrar_movimiento(
          request=self.request,
          accion=f"Vehículo actualizado {self.object.placa}",
          modulo="vehiculos",
        )

        messages.success(self.request, "Vehículo actualizado exitosamente.")
        
        return HttpResponseRedirect(self.get_success_url())



# ====================================================================
# DT5 - DATOS DE TRANSPORTISTAS (Consolidado Conductor + Vehículo)
# ====================================================================


class DT5DetailView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = "taxis/dt5_detail.html"

    def test_func(self):
        return self.request.user.is_authenticated

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        conductor_id = kwargs.get("conductor_id")
        vehicle_id = kwargs.get("vehicle_id")

        conductor = get_object_or_404(
            Conductor.objects.select_related("ubicacion").prefetch_related("vehiculos"),
            pk=conductor_id
        )

        if vehicle_id:
            vehiculo = get_object_or_404(conductor.vehiculos.all(), pk=vehicle_id)
        else:
            vehiculo = conductor.vehiculos.order_by("id").first()

        context["conductor"] = conductor
        context["vehiculo"] = vehiculo
        context["sin_vehiculo"] = vehiculo is None

        # FOTO CONDUCTOR
        avatar = getattr(conductor, "avatar", None)
        context["has_conductor_avatar"] = bool(avatar)
        context["conductor_avatar_url"] = avatar.url if avatar and hasattr(avatar, "url") else None

        # Back URL
        from_module = self.request.GET.get("from", "conductores")
        url_mapping = {
            "vehiculos": reverse("taxis:vehiculo_list"),
            "conductores": reverse("taxis:conductor_list"),
            "conductor_create": reverse("taxis:conductor_list"),
        }
        base_url = url_mapping.get(from_module, reverse("taxis:conductor_list"))

        query_params = self.request.GET.copy()
        query_params.pop("page", None)
        query_params.pop("from", None)
        context["back_url"] = f"{base_url}?{query_params.urlencode()}" if query_params else base_url

        # Edad / fechas
        hoy = timezone.now().date()
        context["today"] = hoy

        if conductor.fechanacimiento:
            edad = hoy.year - conductor.fechanacimiento.year
            if (hoy.month, hoy.day) < (conductor.fechanacimiento.month, conductor.fechanacimiento.day):
                edad -= 1
            context["edad"] = edad

        # Vencimientos conductor
        context["conductor_cedula_vencida"] = bool(conductor.cedula_vencimiento and conductor.cedula_vencimiento < hoy)
        context["conductor_rif_vencido"] = bool(conductor.rif_vencimiento and conductor.rif_vencimiento < hoy)

        # Vencimientos vehículo
        if vehiculo:
            context["cert_medico_vencido"] = bool(vehiculo.medico_vencimiento and vehiculo.medico_vencimiento < hoy)
            context["patente_vencida"] = bool(vehiculo.patente_vencimiento and vehiculo.patente_vencimiento < hoy)
            context["lic_circulacion_vencida"] = bool(vehiculo.licencia_vencimiento and vehiculo.licencia_vencimiento < hoy)
            context["seguro_vencido"] = bool(vehiculo.rcv_vencimiento and vehiculo.rcv_vencimiento < hoy)
        else:
            context["cert_medico_vencido"] = False
            context["patente_vencida"] = False
            context["lic_circulacion_vencida"] = False
            context["seguro_vencido"] = False

        return context


# ====================================================================
# REPORTES DT5 (PDF y EXCEL)
# ====================================================================

def obtener_transportistas_filtrados(request):
    """
    Filtra transportistas (SOLO conductores CON vehículos) aplicando los mismos filtros que DT5ListView
    Respeta: búsqueda, estado y MANTIENE el ID del conductor en orden
    """
    # IMPORTANTE: Solo vehículos (excluye conductores sin vehículo)
    qs = Vehiculo.objects.select_related('conductor').order_by('conductor__id', 'numero_casco')
    
    # Filtro de búsqueda (mismo que en la lista)
    q = request.GET.get('q')
    if q:
        qs = qs.filter(
            Q(numero_casco__icontains=q) |
            Q(placa__icontains=q) |
            Q(conductor__nombres__icontains=q) |
            Q(conductor__apellidos__icontains=q) |
            Q(conductor__cedula_identidad__icontains=q)
        )
    
    # Filtro de estado operativo/inoperativo (mismo que en la lista)
    estado = request.GET.get('estado')
    if estado in ['operativo', 'inoperativo']:
        qs = qs.filter(condicion=estado)
    
    return qs

def obtener_dt5_unificado(request):
    """
    Devuelve filas tipo DT5 basadas en Conductor, con campos de vehículo opcionales.
    
    Parámetros GET:
    - modo=vehiculos (default): solo conductores con vehículo
    - modo=sinvehiculo: solo conductores sin vehículo
    - estado=operativo/inoperativo: filtra condición del vehículo (desde vehiculo_list)
    - estatus=operativo/inoperativo/sinvehiculo: filtra condición (desde conductor_list)
    - genero=M/F: filtra por sexo del conductor
    - q: búsqueda por nombre, apellido, cédula, placa, casco
    """
    # Leer parámetros
    modo = (request.GET.get("modo") or "vehiculos").strip().lower()
    q = (request.GET.get("q") or "").strip()
    
    # Acepta "estado" (vehiculo_list) O "estatus" (conductor_list)
    estado_param = (request.GET.get("estado") or request.GET.get("estatus") or "").strip().lower()
    
    # NUEVO: filtro de género
    genero = (request.GET.get("genero") or "").strip().upper()

    # Subquery: tomar primer vehículo del conductor (ordenado por id)
    vqs = Vehiculo.objects.filter(conductor=OuterRef("pk")).order_by("id")

    qs = Conductor.objects.all().order_by("id").annotate(
        v_marca=Subquery(vqs.values("marca")[:1]),
        v_modelo=Subquery(vqs.values("modelo")[:1]),
        v_anio=Subquery(vqs.values("anio")[:1]),
        v_placa=Subquery(vqs.values("placa")[:1]),
        v_color=Subquery(vqs.values("color")[:1]),
        v_numero_casco=Subquery(vqs.values("numero_casco")[:1]),
        v_capacidad=Subquery(vqs.values("capacidad")[:1]),
        v_bateria_amperaje=Subquery(vqs.values("bateria_amperaje")[:1]),
        v_aceite_viscosidad=Subquery(vqs.values("aceite_viscosidad")[:1]),
        v_cauchos_medida=Subquery(vqs.values("cauchos_medida")[:1]),
        v_diametro_rin=Subquery(vqs.values("diametro_rin")[:1]),
        v_combustible_tipo=Subquery(vqs.values("combustible_tipo")[:1]),
        v_combustible_litros=Subquery(vqs.values("combustible_litros")[:1]),
        v_condicion=Subquery(vqs.values("condicion")[:1]),
    )

    # BÚSQUEDA
    if q:
        qs = qs.filter(
            Q(nombres__icontains=q) |
            Q(apellidos__icontains=q) |
            Q(cedula_identidad__icontains=q) |
            Q(vehiculos__placa__icontains=q) |
            Q(vehiculos__numero_casco__icontains=q)
        ).distinct()

    # FILTRO DE GÉNERO
    if genero in ("M", "F"):
        qs = qs.filter(sexo=genero)

    # MODO/ESTATUS (con/sin vehículo)
    # Si viene "estatus=sinvehiculo" desde conductor_list, forzar modo=sinvehiculo
    if estado_param == "sinvehiculo":
        modo = "sinvehiculo"
    
    if modo == "sinvehiculo":
        qs = qs.filter(vehiculos__isnull=True)
    else:
        qs = qs.filter(vehiculos__isnull=False)
        
        # ESTADO DE VEHÍCULO (operativo/inoperativo) - solo si modo=vehiculos
        if estado_param in ("operativo", "inoperativo"):
            qs = qs.annotate(v_condicion_lower=Lower("v_condicion")).filter(v_condicion_lower=estado_param)

    return qs


@login_required
def reporte_dt5_pdf(request, *args, **kwargs):
    """
    Genera reporte DT5 oficial en PDF con filtros idénticos a la lista.
    Soporta modo=vehiculos (default) y modo=sinvehiculo.
    """
    modulo_origen = kwargs.get("modulo_origen") or "vehiculos"

    filas = obtener_dt5_unificado(request)

    # 🔥 RUTAS DE LOS LOGOS - CORREGIDAS
    logo_transporte, logo_mision = obtener_rutas_logos()

    context = {
        'filas': filas,
        'coop': DATOS_COOP,
        'fecha_hoy': timezone.now(),
        'logo_transporte': logo_transporte,
        'logo_mision': logo_mision,
        'filtros': {
            'q': request.GET.get('q'),
            'estado': request.GET.get('estado'),
            'modo': request.GET.get('modo'),
        }
    }

    template = get_template('taxis/reportes/pdf_dt5.html')
    html = template.render(context)

    result = BytesIO()

    pdf = pisa.pisaDocument(
        BytesIO(html.encode("UTF-8")),
        result,
        encoding='UTF-8'
    )

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        filename = f"DT5_Transportistas_{timezone.now().strftime('%d-%m-%Y')}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        
        # AUDITORÍA
        # (ELIMINA) from taxis.utils import texto_filtros
        registrar_movimiento(
         request,
         accion="exportar_pdf",
         modulo=modulo_origen,
         objeto_tipo="Reporte",
         objeto_nombre="DT5 PDF",
         descripcion=texto_filtros(request.GET, fecha=timezone.localtime()),
        )
 
        
        return response

    return HttpResponse("Error generando PDF", status=500)


@login_required
def reporte_dt5_excel(request, *args, **kwargs):
    """
    Genera reporte DT5 en Excel con estructura completa.
    Soporta modo=vehiculos (default) y modo=sinvehiculo.
    """
    modulo_origen = kwargs.get("modulo_origen") or "vehiculos"
    filas = obtener_dt5_unificado(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DT5 Transportistas"

    # TÍTULO PRINCIPAL
    ws.merge_cells('A1:T1')
    titulo = f"DT5 - DATOS DE TRANSPORTISTAS - {DATOS_COOP['nombre']}"
    
    modo = (request.GET.get('modo') or 'vehiculos').strip().lower()
    estado_filtro = request.GET.get('estado')
    
    if modo == 'sinvehiculo':
        titulo += " (SIN VEHÍCULO)"
    elif estado_filtro == 'operativo':
        titulo += " (SOLO OPERATIVOS)"
    elif estado_filtro == 'inoperativo':
        titulo += " (SOLO INOPERATIVOS)"
    elif request.GET.get('q'):
        titulo += " (FILTRADO)"

    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # ENCABEZADOS PRINCIPALES (Fila 3)
    headers_main = [
        "N°", "NOMBRES", "APELLIDOS", "CÉDULA", "TELÉFONO",
        "MARCA", "MODELO", "AÑO", "COLOR", "PLACA", "CASCO", "CAPACIDAD",
        "BATERÍA", "ACEITE", "COMBUSTIBLE", "CANTIDAD", 
        "CAUCHOS", "RÍN", "OPERATIVO", "INOPERATIVO"
    ]
    ws.append([])  # Fila 2 vacía
    ws.append(headers_main)

    # ESTILOS DE ENCABEZADO
    header_fill = PatternFill(start_color="00b300", end_color="00b300", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # DATOS DE TRANSPORTISTAS
    for i, f in enumerate(filas, start=1):
        cedula_full = f"{f.cedula_prefijo}-{f.cedula_identidad}"
        
        # Checkmarks para campos booleanos
        operativo = "✓" if (f.v_condicion or "").lower() == "operativo" else "—"
        inoperativo = "✓" if (f.v_condicion or "").lower() == "inoperativo" else "—"

        row = [
            i,
            f.nombres or "—",
            f.apellidos or "—",
            cedula_full,
            f.telefono_principal or "—",
            f.v_marca or "—",
            f.v_modelo or "—",
            f.v_anio or "—",
            f.v_color or "—",
            f.v_placa or "—",
            f.v_numero_casco or "—",
            f.v_capacidad or "—",
            f.v_bateria_amperaje or "—",
            f.v_aceite_viscosidad or "—",
            f.v_combustible_tipo or "—",
            f.v_combustible_litros or "—",
            f.v_cauchos_medida or "—",
            f.v_diametro_rin or "—",
            operativo,
            inoperativo
        ]
        ws.append(row)

    # AJUSTAR ANCHOS DE COLUMNA
    dims = {
        'A': 5, 'B': 18, 'C': 18, 'D': 12, 'E': 14, 
        'F': 12, 'G': 12, 'H': 6, 'I': 10, 'J': 12, 
        'K': 10, 'L': 10, 'M': 10, 'N': 10, 'O': 12, 
        'P': 10, 'Q': 10, 'R': 8, 'S': 10, 'T': 12
    }
    for col, width in dims.items():
        ws.column_dimensions[col].width = width

    # GENERAR RESPUESTA EXCEL
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"DT5_Transportistas_{timezone.now().strftime('%d-%m-%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    
    
    # AUDITORÍA (EXCEL)
    registrar_movimiento(
       request,
       accion="exportar_excel",
       modulo=modulo_origen,
       objeto_tipo="Reporte",
       objeto_nombre="DT5 Excel",
       descripcion=texto_filtros(request.GET, fecha=timezone.localtime()),
    )

    return response


   
# ====================================================================
# FINANZAS
# ====================================================================


class DeudaListView(LoginRequiredMixin, ListView):
    model = Deuda
    template_name = "taxis/finanzas/deuda_list.html"
    context_object_name = "deudas"
    paginate_by = 20
    ordering = ['-anio', '-mes']


class PagoListView(LoginRequiredMixin, ListView):
    model = Pago
    template_name = "taxis/finanzas/pago_list.html"
    context_object_name = "pagos"
    ordering = ['-fecha_pago']


# --- VISTAS DEL MÓDULO DE FINANZAS SIMPLIFICADO ---


@login_required
def finanzas_principal(request):
    """Vista principal: Muestra afiliados Pendientes (Deuda Acumulada) vs Solvencias del Mes"""
    hoy = timezone.localtime().date()
    mes_actual = hoy.month
    anio_actual = hoy.year

    config = ConfiguracionFinanzas.get_solo()

    # Búsqueda
    q = request.GET.get('q', '')

    # 1. Filtrar solo conductores ACTIVOS con VEHÍCULO OPERATIVO
    conductores = Conductor.objects.filter(
        estado='activo',
        vehiculos__condicion='operativo'
    ).distinct().select_related('ubicacion')

    if q:
        conductores = conductores.filter(
            Q(cedula_identidad__icontains=q) |
            Q(nombres__icontains=q) |
            Q(apellidos__icontains=q) |
            Q(vehiculos__numero_casco__icontains=q)
        ).distinct()

    # Pre-fetch de pagos para el año actual
    pagos_anio = PagoMensual.objects.filter(
        anio=anio_actual,
        conductor__in=conductores,
        archivado=False  # ✅ FILTRAR NO ARCHIVADOS
    ).values('conductor_id', 'mes')

    # Diccionario de pagos por conductor
    mapa_pagos = {}
    for p in pagos_anio:
        cid = p['conductor_id']
        if cid not in mapa_pagos:
            mapa_pagos[cid] = set()
        mapa_pagos[cid].add(p['mes'])

    pendientes = []
    pagados_completo = []

    conductores = conductores.prefetch_related('vehiculos')

    import calendar

    # Día de vencimiento (por defecto 5)
    dia_venc = getattr(config, 'dia_vencimiento', 5) or 5

    # Ajustar vencimiento al último día del mes actual (por si config pone 31 y el mes tiene 28/30)
    ultimo_dia_mes_actual = calendar.monthrange(anio_actual, mes_actual)[1]
    dia_venc_efectivo = min(dia_venc, ultimo_dia_mes_actual)

    # Si hoy es el día de vencimiento (00:00 ya cuenta) o después, contamos el mes actual como deuda
    tope = mes_actual + 1 if hoy.day >= dia_venc_efectivo else mes_actual
    meses_a_verificar = list(range(1, tope))

    for c in conductores:
        pagos_c = mapa_pagos.get(c.id, set())

        # Verificar Solvencia del Mes Actual (solo para la lista de pagados del mes)
        if mes_actual in pagos_c:
            pago_obj = PagoMensual.objects.filter(
                conductor=c,
                mes=mes_actual,
                anio=anio_actual,
                archivado=False
            ).first()
            pagados_completo.append({'conductor': c, 'pago': pago_obj})

        # Calcular Deuda Acumulada (hasta el tope definido por vencimiento)
        deuda_meses = []
        for m in meses_a_verificar:
            if m not in pagos_c:
                deuda_meses.append(m)

        # IMPORTANTE: si debe meses anteriores, debe aparecer en pendientes aunque ya pagó el mes actual
        if deuda_meses:
            total_deuda = len(deuda_meses) * config.monto_cuota_usd
            pendientes.append({
                'conductor': c,
                'meses_deuda': deuda_meses,
                'total_deuda': total_deuda,
                'cant_meses': len(deuda_meses)
            })

    # Convertir números de mes a nombres
    MESES_NOMBRES = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]

    for item in pendientes:
        item['meses_nombres'] = [MESES_NOMBRES[m - 1] for m in item['meses_deuda']]

    # Ordenar listas
    pendientes.sort(key=lambda x: x['conductor'].id)
    pagados_completo.sort(key=lambda x: x['conductor'].id)

    # Limitar pagados visibles
    LIMITE_INICIAL = 5
    pagados_visibles = pagados_completo[:LIMITE_INICIAL]
    pagados_ocultos = pagados_completo[LIMITE_INICIAL:]
    total_ocultos = len(pagados_ocultos)

    context = {
        'pendientes': pendientes,
        'pagados': pagados_visibles,
        'pagados_ocultos': pagados_ocultos,
        'total_ocultos': total_ocultos,
        'config': config,
        'mes_actual': mes_actual,
        'anio_actual': anio_actual,
        'hoy': hoy,
        'query': q,
        'mes_nombre': PagoMensual(mes=mes_actual).get_nombre_mes()
    }
    return render(request, 'taxis/finanzas/principal.html', context)


@login_required
def finanzas_registrar_pago(request, conductor_id):
    """Registrar pago mensual para un conductor"""
    conductor = get_object_or_404(Conductor, pk=conductor_id)

    vehiculo = conductor.vehiculos.filter(condicion='operativo').first()

    if not vehiculo:
        messages.error(request, 'El conductor no tiene un vehículo operativo asignado.')
        return redirect('taxis:finanzas_principal')

    config = ConfiguracionFinanzas.get_solo()
    hoy = timezone.localtime().date()
    mes_actual = hoy.month
    anio_actual = hoy.year

    # Obtener pagos realizados
    pagos_realizados = PagoMensual.objects.filter(
        conductor=conductor,
        anio=anio_actual,
        archivado=False
    ).values_list('mes', flat=True)

    meses_pagados = set(pagos_realizados)

    MESES_NOMBRES = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]

    import calendar
    from datetime import date

    # Día de vencimiento (por defecto 5)
    dia_venc = getattr(config, 'dia_vencimiento', 5) or 5

    # Ajustar vencimiento al último día del mes actual (por si config pone 31 y el mes tiene 28/30)
    ultimo_dia_mes_actual = calendar.monthrange(anio_actual, mes_actual)[1]
    dia_venc_efectivo = min(dia_venc, ultimo_dia_mes_actual)

    # Mes actual cuenta como adeudado desde las 00:00 del día de vencimiento (>=)
    tope = mes_actual + 1 if hoy.day >= dia_venc_efectivo else mes_actual

    # Meses adeudados
    meses_adeudados = []
    for m in range(1, tope):
        if m not in meses_pagados:
            meses_adeudados.append({
                'mes': m,
                'anio': anio_actual,
                'nombre': MESES_NOMBRES[m - 1],
                'tipo': 'adeudado',
                'es_actual': (m == mes_actual)
            })

    # Meses futuros (incluye el mes actual como "futuro" solo antes del vencimiento)
    meses_futuros = []
    inicio_futuros = mes_actual if hoy.day < dia_venc_efectivo else (mes_actual + 1)
    for m in range(inicio_futuros, 13):
        if m not in meses_pagados:
            meses_futuros.append({
                'mes': m,
                'anio': anio_actual,
                'nombre': MESES_NOMBRES[m - 1],
                'tipo': 'futuro'
            })

    esta_al_dia = len(meses_adeudados) == 0
    deuda_total = len(meses_adeudados) * config.monto_cuota_usd

    # Mes de referencia para vencimiento:
    # - si hay deuda: primer mes adeudado
    # - si no hay deuda y NO ha pagado el mes actual: mes actual (por vencer)
    # - si no hay deuda y ya pagó: próximo mes no pagado (si existe), si no: mes actual
    if meses_adeudados:
        mes_ref = meses_adeudados[0]['mes']
        anio_ref = anio_actual
    else:
        if mes_actual not in meses_pagados:
            mes_ref = mes_actual
            anio_ref = anio_actual
        else:
            prox = None
            for item in meses_futuros:
                prox = item
                break
            if prox:
                mes_ref = prox['mes']
                anio_ref = prox['anio']
            else:
                mes_ref = mes_actual
                anio_ref = anio_actual

    # Ajuste para meses con menos días (ej: febrero)
    ultimo_dia = calendar.monthrange(anio_ref, mes_ref)[1]
    dia_venc_ajustado = min(dia_venc, ultimo_dia)

    fecha_vencimiento = date(anio_ref, mes_ref, dia_venc_ajustado)
    mes_vencimiento_nombre = MESES_NOMBRES[mes_ref - 1]

    # PROCESAR POST
    if request.method == 'POST':
        meses_pagar = request.POST.getlist('meses_pagar[]')

        if not meses_pagar:
            messages.error(request, 'Debes seleccionar al menos un mes para pagar.')
            return redirect('taxis:finanzas_registrar_pago', conductor_id=conductor_id)

        form = PagoMensualForm(request.POST, request.FILES)

        if form.is_valid():
            pagos_registrados = []
            meses_duplicados = []

            for mes_str in meses_pagar:
                mes, anio = map(int, mes_str.split('-'))

                # Verificar duplicados
                if PagoMensual.objects.filter(
                    conductor=conductor,
                    mes=mes,
                    anio=anio,
                    archivado=False
                ).exists():
                    meses_duplicados.append(MESES_NOMBRES[mes - 1])
                    continue

                # Crear el pago
                PagoMensual.objects.create(
                    conductor=conductor,
                    vehiculo=vehiculo,
                    mes=mes,
                    anio=anio,
                    monto_usd=config.monto_cuota_usd,
                    fecha_pago=form.cleaned_data['fecha_pago'],
                    comprobante=form.cleaned_data.get('comprobante'),
                    notas=form.cleaned_data.get('notas', ''),
                    registrado_por=request.user,
                    conductor_nombre_cache=f"{conductor.nombres} {conductor.apellidos}",
                    conductor_cedula_cache=f"{conductor.cedula_prefijo}-{conductor.cedula_identidad}"
                )
                pagos_registrados.append(MESES_NOMBRES[mes - 1])

                # LOG AUDITORÍA - Pago registrado
                log_pago(
                    conductor=conductor,
                    monto=config.monto_cuota_usd,
                    mes=mes,
                    anio=anio,
                    request=request
                )

            # Mensajes
            if meses_duplicados and not pagos_registrados:
                messages.warning(
                    request,
                    f'⚠️ Los meses seleccionados ya están registrados: {", ".join(meses_duplicados)}'
                )
                return redirect('taxis:finanzas_principal')

            if meses_duplicados and pagos_registrados:
                messages.warning(
                    request,
                    f'⚠️ Ya registrados: {", ".join(meses_duplicados)}'
                )

            if pagos_registrados:
                MovimientoAudit.objects.create(
                    usuario=request.user,
                    accion=f'Registró pagos de {conductor}: {", ".join(pagos_registrados)}',
                    modulo='Finanzas'
                )
                messages.success(request, f'✅ Pagos registrados: {", ".join(pagos_registrados)}')

            return redirect('taxis:finanzas_principal')

        else:
            # Mostrar errores del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')

    else:
        # GET: Formulario vacío
        form = PagoMensualForm()

    context = {
        'conductor': conductor,
        'vehiculo': vehiculo,
        'form': form,
        'config': config,
        'hoy': hoy,
        'fecha_vencimiento': fecha_vencimiento,
        'mes_vencimiento_nombre': mes_vencimiento_nombre,
        'meses_adeudados': meses_adeudados,
        'meses_futuros': meses_futuros,
        'esta_al_dia': esta_al_dia,
        'deuda_total': deuda_total,
        'anio_actual': anio_actual,
    }
    return render(request, 'taxis/finanzas/registrar_pago_v2.html', context)


@login_required
def finanzas_historial(request):
    """Historial financiero con estadísticas y paginación"""
    from datetime import date
    import calendar

    # Verificación de permisos
    if not hasattr(request.user, 'role'):
        messages.error(request, '⚠️ Tu usuario no tiene un rol asignado.')
        return redirect('taxis:panel_general')

    if request.user.role.upper() != 'PRESIDENTE':
        messages.warning(request, '⚠️ Solo el presidente puede acceder al historial financiero.')
        return redirect('taxis:finanzas_principal')

    hoy = date.today()
    anio_seleccionado = int(request.GET.get('anio', hoy.year))
    mes_seleccionado = request.GET.get('mes', '')
    query = request.GET.get('q', '').strip()
    orden = request.GET.get('orden', '-anio,-mes,-fecha_pago')

    # Día de vencimiento (para puntualidad)
    config = ConfiguracionFinanzas.get_solo()
    dia_venc = getattr(config, 'dia_vencimiento', 5) or 5

    # Consulta base - FILTRAR NO ARCHIVADOS
    pagos_qs = PagoMensual.objects.filter(
        archivado=False
    ).select_related('conductor', 'vehiculo').filter(anio=anio_seleccionado)

    # Filtro por mes
    if mes_seleccionado:
        pagos_qs = pagos_qs.filter(mes=int(mes_seleccionado))

    # Búsqueda
    if query:
        pagos_qs = pagos_qs.filter(
            Q(conductor__nombres__icontains=query) |
            Q(conductor__apellidos__icontains=query) |
            Q(conductor__cedula_identidad__icontains=query) |
            Q(vehiculo__numero_casco__icontains=query)
        ).distinct()

    # Ordenamiento seguro
    orden = request.GET.get("orden", "-anio,-mes,-fecha_pago")
    orden_lista = [o.strip() for o in orden.split(",") if o.strip()]
    permitidos = {'anio', '-anio', 'mes', '-mes', 'fecha_pago', '-fecha_pago'}
    orden_final = [o for o in orden_lista if o in permitidos] or ['-anio', '-mes', '-fecha_pago']
    pagos_qs = pagos_qs.order_by(*orden_final)

    # Estadísticas globales
    stats_globales = pagos_qs.aggregate(
        total_recaudado=Sum('monto_usd'),
        total_pagos=Count('id'),
        afiliados_pagaron=Count('conductor', distinct=True)
    )

    # Calcular promedio
    if stats_globales['total_pagos'] and stats_globales['total_pagos'] > 0:
        stats_globales['promedio_pago'] = stats_globales['total_recaudado'] / stats_globales['total_pagos']
    else:
        stats_globales['promedio_pago'] = 0

    # Manejar valores None
    stats_globales['total_recaudado'] = stats_globales['total_recaudado'] or 0
    stats_globales['total_pagos'] = stats_globales['total_pagos'] or 0
    stats_globales['afiliados_pagaron'] = stats_globales['afiliados_pagaron'] or 0

    # Paginación
    paginator = Paginator(pagos_qs, 20)
    page = request.GET.get('page', 1)

    try:
        pagos_paginados = paginator.page(page)
    except PageNotAnInteger:
        pagos_paginados = paginator.page(1)
    except EmptyPage:
        pagos_paginados = paginator.page(paginator.num_pages)

    # Enriquecer pagos con estados
    for pago in pagos_paginados:
        ultimo_dia_pago = calendar.monthrange(pago.anio, pago.mes)[1]
        dia_venc_pago = min(dia_venc, ultimo_dia_pago)

        if pago.anio > hoy.year:
            pago.es_adelantado = True
            pago.es_puntual = False
        elif pago.anio == hoy.year and pago.mes > hoy.month:
            pago.es_adelantado = True
            pago.es_puntual = False
        elif pago.anio == hoy.year and pago.mes == hoy.month and pago.fecha_pago.day <= dia_venc_pago:
            pago.es_puntual = True
            pago.es_adelantado = False
        else:
            pago.es_adelantado = False
            pago.es_puntual = False

    # Estadísticas del filtro actual
    total_filtrado = pagos_qs.aggregate(total=Sum('monto_usd'))['total'] or 0

    stats_filtradas = {
        'total_pagos_filtrados': pagos_paginados.paginator.count,
        'total_filtrado': total_filtrado,
        'afiliados_unicos': pagos_qs.values('conductor').distinct().count()
    }

    # Años disponibles
    anios_disponibles = list(range(hoy.year, hoy.year - 5, -1))

    # Meses
    meses_lista = [
        {'num': 1, 'nombre': 'Enero'},
        {'num': 2, 'nombre': 'Febrero'},
        {'num': 3, 'nombre': 'Marzo'},
        {'num': 4, 'nombre': 'Abril'},
        {'num': 5, 'nombre': 'Mayo'},
        {'num': 6, 'nombre': 'Junio'},
        {'num': 7, 'nombre': 'Julio'},
        {'num': 8, 'nombre': 'Agosto'},
        {'num': 9, 'nombre': 'Septiembre'},
        {'num': 10, 'nombre': 'Octubre'},
        {'num': 11, 'nombre': 'Noviembre'},
        {'num': 12, 'nombre': 'Diciembre'},
    ]

    context = {
        'pagos': pagos_paginados,
        'stats': {**stats_globales, **stats_filtradas},
        'anio_seleccionado': anio_seleccionado,
        'mes_seleccionado': int(mes_seleccionado) if mes_seleccionado else '',
        'query': query,
        'orden': ",".join(orden_final),
        'anios_disponibles': anios_disponibles,
        'meses_lista': meses_lista,
    }

    return render(request, 'taxis/finanzas/historial.html', context)


@login_required
def finanzas_ver_pago(request, pago_id):
    """Detalle de un pago mensual con agrupación inteligente"""
    pago = get_object_or_404(
        PagoMensual.objects.select_related('conductor', 'vehiculo', 'registrado_por'),
        id=pago_id
    )

    # Detectar pagos relacionados (misma transacción)
    pagos_agrupados = PagoMensual.objects.filter(
        conductor=pago.conductor,
        fecha_pago=pago.fecha_pago,
        archivado=False
    ).order_by('anio', 'mes')

    # Si hay comprobante, filtrar por mismo comprobante
    if pago.comprobante:
        pagos_agrupados = pagos_agrupados.filter(comprobante=pago.comprobante.name)

    # Calcular total acumulado
    total_acumulado = sum(p.monto_usd for p in pagos_agrupados)

    # Generar string de meses
    MESES_NOMBRES = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]

    meses_str = ', '.join([MESES_NOMBRES[p.mes - 1] for p in pagos_agrupados])

    # Determinar si es pago múltiple
    es_pago_multiple = pagos_agrupados.count() > 1

    context = {
        'pago': pago,
        'pagos_agrupados': pagos_agrupados,
        'total_acumulado': total_acumulado,
        'meses_str': meses_str,
        'es_pago_multiple': es_pago_multiple,
        'coop': DATOS_COOP,
    }

    # AUDITORÍA (AQUÍ)
    MovimientoAudit.objects.create(
        usuario=request.user,
        accion="Imprimió ticket de pago",
        modulo="Finanzas",
        descripcion=f"Pago ID: {pago_id}"
    )

    return render(request, 'taxis/finanzas/ver_pago.html', context)


# Sistema antiguo de deudas (compatibilidad)
class RegistrarPagoView(LoginRequiredMixin, CreateView):
    model = Pago
    form_class = PagoForm
    template_name = "taxis/finanzas/pago_form.html"
    success_url = reverse_lazy("taxis:deuda_list")

    def dispatch(self, request, *args, **kwargs):
        self.deuda = get_object_or_404(Deuda, pk=self.kwargs['pk'])
        if self.deuda.pagada:
            messages.warning(request, "Deuda ya pagada.")
            return redirect("taxis:deuda_list")
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        pago = form.save(commit=False)
        pago.deuda = self.deuda
        pago.save()
        self.deuda.pagada = True
        self.deuda.save()

        MovimientoAudit.objects.create(
            usuario=self.request.user,
            accion=f"Registró pago de {self.deuda.conductor.nombres}",
            modulo="Finanzas"
        )
        messages.success(self.request, "Pago registrado correctamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['deuda'] = self.deuda
        return context


# ====================================================================
# Filtro de los Vehículos
# ====================================================================

def obtener_vehiculos_filtrados(request):
    """Filtra vehículos por estado (operativo/inoperativo) y búsqueda."""
    qs = Vehiculo.objects.select_related('conductor').order_by('numero_casco')
    
    q = request.GET.get('q')
    if q:
        qs = qs.filter(
            Q(numero_casco__icontains=q) |
            Q(placa__icontains=q) |
            Q(conductor__nombres__icontains=q) |
            Q(conductor__apellidos__icontains=q) |
            Q(conductor__cedula_identidad__icontains=q)
        )
    
    estado = request.GET.get('estado')
    if estado in ['operativo', 'inoperativo']:
        qs = qs.filter(condicion=estado)
    
    return qs


# ====================================================================
# AUDITORÍA Y NOTIFICACIONES
# ====================================================================

class MovimientoAuditListView(LoginRequiredMixin, ListView):
    model = MovimientoAudit
    template_name = "taxis/movimientoaudit_list.html"
    context_object_name = "movimientos"
    paginate_by = 30

    # Keys (BD / filtros) -> Label (UI)
    MODULO_MAP = {
        "autenticacion": "Autenticación",
        "afiliados": "Afiliados",
        "vehiculos": "Vehículos",
        "finanzas": "Finanzas",
        "perfiles": "Perfiles",
        "auditoria": "Auditoría",
    }

    ORDERING_MAP = {
        "recientes": "-fecha",
        "antiguos": "fecha",
        "modulo_az": "modulo",
        "modulo_za": "-modulo",
    }

    def get_ordering(self):
        orden = (self.request.GET.get("orden") or "").strip()
        return self.ORDERING_MAP.get(orden, "-fecha")

    def get_queryset(self):
        qs = MovimientoAudit.objects.select_related("usuario").all()

        # =====================
        # Búsqueda general (q)
        # =====================
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(usuario__first_name__icontains=q) |
                Q(usuario__last_name__icontains=q) |
                Q(usuario__email__icontains=q) |
                Q(descripcion__icontains=q) |
                Q(accion__icontains=q) |
                Q(modulo__icontains=q)
            )

        # ==========
        # Módulo (Opción B: EXACT MATCH, solo claves)
        # ==========
        modulo = (self.request.GET.get("modulo") or "").strip().lower()
        if modulo:
            # Si te llega cualquier cosa rara por URL, no filtramos para evitar "0 resultados" confusos
            if modulo in self.MODULO_MAP:
                qs = qs.filter(modulo=modulo)

        # ======================
        # Período
        # ======================
        periodo = (self.request.GET.get("periodo") or "").strip().lower()
        if periodo:
            now = timezone.localtime(timezone.now())

            if periodo == "hoy":
                qs = qs.filter(fecha__date=now.date())

            elif periodo == "7d":
                qs = qs.filter(fecha__gte=now - timedelta(days=7))

            elif periodo == "mes":
                inicio_mes = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                qs = qs.filter(fecha__gte=inicio_mes)

            elif periodo == "anio":
                inicio_anio = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
                qs = qs.filter(fecha__gte=inicio_anio)

        # ==================================
        # Año/Mes (opcional, si lo mantienes)
        # ==================================
        anio = (self.request.GET.get("anio") or "").strip()
        mes = (self.request.GET.get("mes") or "").strip()

        if anio.isdigit():
            qs = qs.filter(fecha__year=int(anio))

        if mes.isdigit():
            m = int(mes)
            if 1 <= m <= 12:
                qs = qs.filter(fecha__month=m)

        return qs.order_by(self.get_ordering())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context["q"] = self.request.GET.get("q", "")
        context["modulo_filtro"] = self.request.GET.get("modulo", "")
        context["periodo_filtro"] = self.request.GET.get("periodo", "")
        context["anio_filtro"] = self.request.GET.get("anio", "")
        context["mes_filtro"] = self.request.GET.get("mes", "")
        context["orden_filtro"] = self.request.GET.get("orden", "")

        context["modulos_choices"] = list(self.MODULO_MAP.items())

        params = self.request.GET.copy()
        params.pop("page", None)
        context["query_params"] = params.urlencode()

        return context
    
def link_callback(uri, rel):
    # Quitar dominio y querystring si vinieran (por si acaso)
    uri = urlparse(uri).path

    static_url = settings.STATIC_URL or "/static/"
    media_url = settings.MEDIA_URL or "/media/"

    # Normalizar por si en settings está "static/" sin slash inicial
    if not static_url.startswith("/"):
        static_url = "/" + static_url
    if not media_url.startswith("/"):
        media_url = "/" + media_url

    # 1) STATIC: /static/img/...  ->  img/...
    if uri.startswith(static_url):
        relative_path = uri[len(static_url):].lstrip("/")  # "img/logo.png"
        absolute_path = finders.find(relative_path)
        if absolute_path:
            return os.path.realpath(absolute_path)
        return uri

    # 2) MEDIA: /media/... -> MEDIA_ROOT/...
    if uri.startswith(media_url):
        path = os.path.join(settings.MEDIA_ROOT, uri[len(media_url):].lstrip("/"))
        if os.path.isfile(path):
            return os.path.realpath(path)
        return uri

    return uri

@login_required
def auditoria_pdf(request):
    # 1) Reusar EXACTAMENTE los filtros de la lista
    view = MovimientoAuditListView()
    view.request = request
    movimientos = view.get_queryset()

    # 2) Render HTML
    context = {
        "movimientos": movimientos,
        "fecha_generacion": timezone.now(),
    }
    template = get_template("taxis/reportes/auditoria_pdf.html")
    html = template.render(context)

    # 3) Generar PDF
    result = BytesIO()
    pdf = pisa.pisaDocument(
        BytesIO(html.encode("UTF-8")),
        result,
        encoding="UTF-8",
        link_callback=link_callback
    )

    if pdf.err:
        return HttpResponse("Error generando PDF", status=500)

    # 4) Auditar exportación (con filtros y cantidad)
    filtros_txt = texto_filtros({k: v for k, v in request.GET.lists()})

    total = movimientos.count()

    registrar_movimiento(
        request=request,
        accion="exportar_pdf",
        modulo="auditoria",
        objeto_tipo="Reporte",
        objeto_nombre="Log de Auditoría (PDF)",
        descripcion=f"{filtros_txt} | Registros: {total}",
    )

    # 5) Respuesta
    response = HttpResponse(result.getvalue(), content_type="application/pdf")
    filename = f'Auditoria_{timezone.now().strftime("%d-%m-%Y_%H-%M")}.pdf'
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


@login_required
def panel_general(request):
    """Panel principal con tarjetas, alertas y stats"""

    import calendar

    hoy = timezone.localtime().date()

    # ====================================================================
    # 1. CONTADORES TARJETAS
    # ====================================================================
    total_afiliados = Conductor.objects.filter(estado='activo').count()
    total_vehiculos = Vehiculo.objects.count()

    # Pagos pendientes (conductores con vehículo operativo que NO pagaron este mes)
    conductores_con_vehiculo = Conductor.objects.filter(
        estado='activo',
        vehiculos__condicion='operativo'
    ).distinct()

    pagos_realizados = PagoMensual.objects.filter(
        mes=hoy.month,
        anio=hoy.year,
        archivado=False
    ).values_list('conductor_id', flat=True)

    # ✅ Ajuste por vencimiento: contar pendientes solo desde las 00:00 del día de vencimiento
    config_finanzas = ConfiguracionFinanzas.get_solo()
    dia_venc = getattr(config_finanzas, 'dia_vencimiento', 5) or 5

    ultimo_dia_mes = calendar.monthrange(hoy.year, hoy.month)[1]
    dia_venc_efectivo = min(dia_venc, ultimo_dia_mes)

    if hoy.day >= dia_venc_efectivo:
        pagos_pendientes = conductores_con_vehiculo.exclude(id__in=pagos_realizados).count()
    else:
        pagos_pendientes = 0

    legal_docs = CooperativaLegalDocs.get_solo()
    legal_form = CooperativaLegalDocsForm(instance=legal_docs)
    edit_mode = request.GET.get("edit") == "1"

    # ====================================================================
    # 2. INFORMACIÓN DE LA COOPERATIVA (USA DATOS_COOP)
    # ====================================================================
    coop = {
        'nombre': 'WILSON TORRES 33, R.L.',
        'rif': 'J-40126249-0',
        'direccion': 'CALLE LA GLORIA, CASA N° 115-C, SECTOR JESÚS BANDRES. SAN JUAN DE LOS MORROS - ESTADO GUÁRICO',
        'presidente': 'WILSON TORRES',
        'telefono': '0416-6444886',
        'email': 'wilsontorres27@gmail.com',
        'municipio': 'JUAN GERMAN ROSCIO',
    }

    # ====================================================================
    # 3. ALERTAS DE DOCUMENTOS VENCIDOS (CONDUCTORES)
    # ====================================================================
    alertas_documentos = []

    conductores = Conductor.objects.filter(estado='activo').prefetch_related('vehiculos')
    for c in conductores:
        vehiculo_c = c.vehiculos.first()

        # CÉDULA VENCIDA
        if c.cedula_vencimiento and c.cedula_vencimiento <= hoy:
            alertas_documentos.append({
                'tipo': 'conductor',
                'conductor_id': c.id,
                'vehiculo_id': vehiculo_c.id if vehiculo_c else None,
                'conductor_nombre': f"{c.nombres} {c.apellidos}",
                'titulo': 'Cédula vencida',
                'descripcion': f"El documento Cédula venció el {c.cedula_vencimiento.strftime('%d/%m/%Y')}",
                'documento_tipo': 'Cédula',
                'fecha_vencimiento': c.cedula_vencimiento
            })

        # RIF VENCIDO
        if c.rif_vencimiento and c.rif_vencimiento <= hoy:
            alertas_documentos.append({
                'tipo': 'conductor',
                'conductor_id': c.id,
                'vehiculo_id': vehiculo_c.id if vehiculo_c else None,
                'conductor_nombre': f"{c.nombres} {c.apellidos}",
                'titulo': 'RIF vencido',
                'descripcion': f"El documento RIF venció el {c.rif_vencimiento.strftime('%d/%m/%Y')}",
                'documento_tipo': 'RIF',
                'fecha_vencimiento': c.rif_vencimiento
            })

    # ====================================================================
    # 4. ALERTAS DE DOCUMENTOS VENCIDOS (VEHÍCULOS)
    # ====================================================================
    vehiculos = Vehiculo.objects.all()
    for v in vehiculos:
        checks = [
            ('Patente', v.patente_vencimiento),
            ('Licencia', v.licencia_vencimiento),
            ('RCV', v.rcv_vencimiento),
            ('Médico', v.medico_vencimiento)
        ]
        for tipo_doc, fecha in checks:
            if fecha and fecha <= hoy:
                alertas_documentos.append({
                    'tipo': 'vehiculo',
                    'vehiculo_id': v.id,
                    'conductor_id': v.conductor.id,
                    'conductor_nombre': f"{v.conductor.nombres} {v.conductor.apellidos}",
                    'titulo': f'{tipo_doc} vencida' if tipo_doc != 'RCV' and tipo_doc != 'Médico' else f'{tipo_doc} vencido',
                    'descripcion': f"El documento {tipo_doc} del vehículo {v.placa} venció el {fecha.strftime('%d/%m/%Y')}",
                    'documento_tipo': tipo_doc,
                    'fecha_vencimiento': fecha
                })

    # ====================================================================
    # 5. ORDENAR ALERTAS POR FECHA (más recientes primero)
    # ====================================================================
    alertas_documentos.sort(key=lambda x: x['fecha_vencimiento'], reverse=True)

    # ====================================================================
    # 6. CONTEXT COMPLETO
    # ====================================================================
    context = {
        'total_afiliados': total_afiliados,
        'total_vehiculos': total_vehiculos,
        'pagos_pendientes': pagos_pendientes,
        'alertas_documentos': alertas_documentos,
        'coop': coop,
        'today': hoy,
    }

    context.update({
        "legal_docs": legal_docs,   # tu template usa legal_docs
        "legal_form": legal_form,   # tu template usa legal_form
        "edit_mode": edit_mode,
    })

    return render(request, 'taxis/panel_general.html', context)


  
@login_required
def ayuda_sistema(request):
    return render(request, 'taxis/ayuda.html')


@login_required
@require_POST
def actualizar_avatar_presidente(request):
    if 'avatar' in request.FILES:
        request.user.avatar = request.FILES['avatar']
        request.user.save()
        registrar_movimiento(
          request=request,
          accion="Actualizó foto de perfil",
          modulo="perfiles",
        )

        return JsonResponse({'status': 'ok', 'url': request.user.avatar.url})
    return JsonResponse({'status': 'error'}, status=400)

# ====================================================================
# AUTH Y REGISTRO
# ====================================================================

class CustomLoginView(LoginView):
    template_name = "registration/login.html"
    authentication_form = EmailOrUsernameAuthenticationForm

    def get_success_url(self):
        return reverse_lazy("taxis:panel_general")
    
    def form_valid(self, form):
        response = super().form_valid(form)  # aquí ya está logueado
        registrar_movimiento(
            request=self.request,
            accion="login",
            modulo="autenticacion",
            objeto_tipo=self.request.user.__class__.__name__,
            objeto_id=self.request.user.id,
            objeto_nombre=getattr(self.request.user, "username", "") or "",
            descripcion=f"Inicio de sesión: {getattr(self.request.user, 'email', '') or ''}",
            usuario=self.request.user,
        )
        return response


def login_redirect_view(request):
    return redirect("taxis:panel_general")


def select_role(request):
    presidente_activo_existe = CustomUser.objects.filter(role='presidente', is_active=True).exists()

    if request.method == "POST":
        rol_seleccionado = request.POST.get('role')
        if rol_seleccionado == 'presidente':
            if presidente_activo_existe:
                messages.error(request, "Ya existe un presidente activo en el sistema.")
                return redirect('taxis:login')
            return redirect('taxis:register_presidente')

    return render(request, 'taxis/select_role.html', {'presidente_existente': presidente_activo_existe})



def register_presidente(request):
    """
    Registro de Presidente:
    - Crea PendingPresidentRegistration en BD (token UUID del modelo).
    - Envía correo con /confirmar-presidente/<uuid:token>/
    """
    # Si ya hay un pending en sesión, mostrar la pantalla de verificación
    if request.session.get('pending_president_token') and request.session.get('pending_email'):
        return render(request, 'taxis/verification_pending.html', {
            'pending_email': request.session.get('pending_email')
        })

    if request.method == 'POST':
        form = PresidenteRegisterForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data

            # 1) No permitir presidente activo
            if CustomUser.objects.filter(role='presidente', is_active=True).exists():
                messages.error(request, 'Ya existe un presidente activo.')
                return render(request, 'registration/register_presidente.html', {'form': form})

            # 2) No email duplicado en CustomUser
            if CustomUser.objects.filter(email__iexact=cd['email']).exists():
                messages.error(request, 'Email ya registrado.')
                return render(request, 'registration/register_presidente.html', {'form': form})

            # 3) No email duplicado en PendingPresidentRegistration
            if PendingPresidentRegistration.objects.filter(email__iexact=cd['email']).exists():
                messages.error(request, 'Este email ya está en proceso de registro. Revisa tu correo.')
                return render(request, 'registration/register_presidente.html', {'form': form})

            phone_number = (cd.get("phone_number") or "").strip()
            phone_country = (cd.get("phone_country") or "").strip()

            if not phone_number:
                messages.error(request, "El número de teléfono es obligatorio.")
                return render(request, 'registration/register_presidente.html', {'form': form})

            # 4) Crear registro pendiente en BD (token UUIDField del modelo)
            pending = PendingPresidentRegistration.objects.create(
                username=cd['username'],
                first_name=cd['first_name'],
                last_name=cd['last_name'],
                email=cd['email'],
                phone_country=cd.get('phone_country', ''),
                phone_number=cd['phone_number'],
                fecha_nacimiento=cd['fechanacimiento'],
                sexo=cd['sexo'],
                password_hash=make_password(cd['password1']),
                expires_at=timezone.now() + timedelta(hours=24),
            )

            # Guardar datos mínimos en sesión para: pantalla "pendiente", reenviar, cancelar
            request.session['pending_email'] = pending.email
            request.session['pending_president_token'] = str(pending.token)
            request.session.modified = True

            # Enviar correo inicial
            send_president_activation_email_initial(request, pending.email, pending.token)

            messages.success(request, f'¡Registrado! Revisa {pending.email}')
            return redirect('taxis:verification_pending')  # ✅ CAMBIO AQUÍ

        # Form inválido
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f'{field}: {error}')
        return render(request, 'registration/register_presidente.html', {'form': form})

    # GET
    form = PresidenteRegisterForm()
    return render(request, 'registration/register_presidente.html', {'form': form})

@require_POST
@csrf_protect
def cancel_registration(request):
    """Cancelar registro pendiente (borra pending en BD si existe) y limpiar sesión."""
    try:
        token = request.session.get('pending_president_token')

        if token:
            PendingPresidentRegistration.objects.filter(token=token).delete()

        request.session.pop('pending_president_token', None)
        request.session.pop('pending_email', None)
        request.session.modified = True

        return JsonResponse({
            'message': 'Registro cancelado. Redirigiendo...',
            'success': True
        }, status=200)

    except Exception as e:
        print(f"❌ ERROR CANCELANDO REGISTRO: {str(e)}")
        return JsonResponse({
            'error': 'Error al cancelar registro.',
            'success': False
        }, status=500)


def confirm_president_registration(request, token):
    """
    Confirma token (UUID) → valida expiración → crea usuario → login → activation_success
    NOTA: Esta vista debe estar en urls.py como: <uuid:token>
    """
    pending = get_object_or_404(PendingPresidentRegistration, token=token)

    # Expirado
    if timezone.now() > pending.expires_at:
        pending.delete()
        messages.error(request, 'Token expirado. Regístrate de nuevo.')
        return redirect('taxis:register_presidente')

    # Presidente ya existe
    if CustomUser.objects.filter(role='presidente', is_active=True).exists():
        pending.delete()
        messages.error(request, 'Ya existe presidente activo.')
        return redirect('taxis:login')

    # Seguridad: evitar colisiones por email
    if CustomUser.objects.filter(email__iexact=pending.email).exists():
        pending.delete()
        messages.error(request, 'Email ya registrado.')
        return redirect('taxis:register_presidente')

    # Validación teléfono (misma lógica que venías usando)
    phone_raw = (pending.phone_number or '').strip()
    if not phone_raw:
        pending.delete()
        messages.error(request, "Tu registro no tiene teléfono. Regístrate de nuevo.")
        return redirect('taxis:register_presidente')

    telefono_limpio = ''.join(filter(str.isdigit, phone_raw))
    telefono_comparar = telefono_limpio[-10:] if len(telefono_limpio) >= 10 else telefono_limpio

    for u in CustomUser.objects.all():
        tlf_user = getattr(u, 'phone_number', '')
        if tlf_user:
            db_limpio = ''.join(filter(str.isdigit, str(tlf_user)))
            db_comparar = db_limpio[-10:] if len(db_limpio) >= 10 else db_limpio
            if db_comparar == telefono_comparar:
                messages.error(request, 'Teléfono ya registrado.')
                return redirect('taxis:register_presidente')

    for conductor in Conductor.objects.all():
        if conductor.telefono_principal:
            db_limpio = ''.join(filter(str.isdigit, conductor.telefono_principal))
            db_comparar = db_limpio[-10:] if len(db_limpio) >= 10 else db_limpio
            if db_comparar == telefono_comparar:
                messages.error(request, 'Teléfono usado por conductor.')
                return redirect('taxis:register_presidente')

    # Crear usuario FINAL (usa el hash guardado en pending.password_hash)
    user = CustomUser.objects.create(
        username=pending.username,
        first_name=pending.first_name,
        last_name=pending.last_name,
        email=pending.email,
        phone_country=pending.phone_country,
        phone_number=pending.phone_number,
        fecha_nacimiento=pending.fecha_nacimiento,
        sexo=pending.sexo,
        role='presidente',
        is_active=True,
        is_email_verified=True
    )
    user.password = pending.password_hash
    user.save()

    pending.delete()

    request.session.pop('pending_president_token', None)
    request.session.pop('pending_email', None)

    login(request, user)
    messages.success(request, f'¡Bienvenido {user.first_name}!')
    return redirect('taxis:activation_success')


def send_president_activation_email_initial(request, email, token):
    """Envío inicial (token UUID)."""
    url = request.build_absolute_uri(
        reverse('taxis:confirm_president_registration', kwargs={'token': str(token)})
    )

    mensaje_html = format_html("""
        <div style="font-family: Nunito, sans-serif; max-width: 600px; margin: 0 auto;">
            <h2 style="color: #00b300;">¡Activa tu cuenta de Presidente!</h2>
            <p>Haz clic para confirmar:</p>
            <a href="{}" style="display: inline-block; background: #00b300; color: white;
                padding: 15px 30px; text-decoration: none; border-radius: 50px;
                font-weight: 800; font-size: 16px;">Confirmar Cuenta</a>
            <p style="margin-top: 20px; font-size: 14px; color: #666;">
                Expira en 24h. Ignora si no solicitaste.
            </p>
        </div>
    """, url)

    send_mail(
        'Confirma cuenta Presidente - WILSON TORRES 33',
        f'Activa aquí: {url}',
        settings.DEFAULT_FROM_EMAIL or 'no-reply@cooperativa.com',
        [email],
        html_message=mensaje_html,
        fail_silently=False,
    )
    print(f"✅ CORREO INICIAL ENVIADO A: {email}")


def send_president_activation_email(request, email, token):
    """Envío para reenvío (token UUID)."""
    url = request.build_absolute_uri(
        reverse('taxis:confirm_president_registration', kwargs={'token': str(token)})
    )

    mensaje_html = format_html("""
        <div style="font-family: Nunito, sans-serif; max-width: 600px; margin: 0 auto;">
            <h2 style="color: #00b300;">¡Activa tu cuenta de Presidente!</h2>
            <p>Haz clic para confirmar:</p>
            <a href="{}" style="display: inline-block; background: #00b300; color: white;
                padding: 15px 30px; text-decoration: none; border-radius: 50px;
                font-weight: 800; font-size: 16px;">Confirmar Cuenta</a>
            <p style="margin-top: 20px; font-size: 14px; color: #666;">
                Expira en 24h. Ignora si no solicitaste.
            </p>
        </div>
    """, url)

    send_mail(
        'Confirma cuenta Presidente - WILSON TORRES 33',
        f'Activa aquí: {url}',
        settings.DEFAULT_FROM_EMAIL or 'no-reply@cooperativa.com',
        [email],
        html_message=mensaje_html,
        fail_silently=False,
    )
    print(f"✅ CORREO REENVIADO A: {email}")


@require_POST
@csrf_protect
def activation_resend(request):
    """
    Reenvío con límite (cache):
    - Máximo 3 reenvíos
    - Bloqueo 30 min si se agotan
    El token sale de BD (PendingPresidentRegistration), no de sesión.
    """
    email = (request.session.get('pending_email') or '').strip()
    if not email:
        return JsonResponse({'error': 'Sesión expirada. Por favor, regístrate de nuevo.'}, status=400)

    pending = (PendingPresidentRegistration.objects
               .filter(email__iexact=email)
               .order_by('-created_at')
               .first())

    if not pending:
        return JsonResponse({'error': 'No hay registro pendiente.'}, status=404)

    if timezone.now() > pending.expires_at:
        pending.delete()
        request.session.pop('pending_president_token', None)
        request.session.pop('pending_email', None)
        return JsonResponse({'error': 'Token expirado. Por favor, regístrate de nuevo.'}, status=410)

    cache_key_count = f"activation_resend_count_{email.lower()}"
    cache_key_locked = f"activation_resend_locked_{email.lower()}"

    locked_until = cache.get(cache_key_locked)
    if locked_until:
        remaining_seconds = int((locked_until - timezone.now()).total_seconds())
        if remaining_seconds > 0:
            return JsonResponse({
                'locked': True,
                'retry_after': remaining_seconds,
                'error': f'Demasiados intentos. Intenta en {remaining_seconds // 60}:{remaining_seconds % 60:02d}.'
            }, status=429)
        cache.delete(cache_key_locked)
        cache.delete(cache_key_count)

    resend_count = cache.get(cache_key_count, 0)
    if resend_count >= 3:
        locked_until = timezone.now() + timedelta(minutes=30)
        cache.set(cache_key_locked, locked_until, 1800)
        return JsonResponse({
            'locked': True,
            'retry_after': 1800,
            'error': 'Demasiados intentos. Intenta en 30 minutos.'
        }, status=429)

    try:
        send_president_activation_email(request, pending.email, pending.token)
        cache.set(cache_key_count, resend_count + 1, 86400)  # contador dura 24h

        remaining = max(0, 3 - resend_count - 1)
        return JsonResponse({'message': 'Correo reenviado correctamente.', 'remaining': remaining}, status=200)

    except Exception as e:
        print(f"❌ ERROR ENVIANDO CORREO: {str(e)}")
        return JsonResponse({'error': 'No se pudo enviar el correo. Intenta más tarde.'}, status=500)


@method_decorator(ensure_csrf_cookie, name='dispatch')
class VerificationPendingView(TemplateView):
    template_name = 'taxis/verification_pending.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pending_email'] = self.request.session.get('pending_email', '')
        return context


class ActivationSuccessView(LoginRequiredMixin, TemplateView):
    template_name = 'taxis/activation_success.html'
    login_url = 'taxis:login'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.role != 'presidente':
            messages.error(request, 'Acceso denegado.')
            return redirect('taxis:login')
        return super().dispatch(request, *args, **kwargs)

    # ====================================================================
# AUTENTICACIÓN - LOGOUT CON AUDITORÍA
# ====================================================================

def logout_view(request):
    """
    Cierra sesión del usuario y registra en auditoría.
    Llamada por: /logout/
    """
    from django.contrib.auth import logout
    from django.contrib import messages
    from taxis.utils import log_logout
    
    # Registrar logout ANTES de cerrar sesión
    log_logout(request)
    
    # Cerrar sesión oficial
    logout(request)
    
    # Mensaje de confirmación
    messages.success(request, '✅ Sesión cerrada correctamente.')
    
    # Redirigir a login
    return redirect('taxis:login')

MAX_RESENDS = 3
LOCK_SECONDS = 60 * 30  # 30 minutos

def _pwreset_count_key(email: str) -> str:
    return f"pwreset:count:{email.lower()}"

def _pwreset_lock_key(email: str) -> str:
    return f"pwreset:lock:{email.lower()}"

@require_POST
@csrf_protect
def password_reset_resend(request):
    """
    Reenviar correo de reset de contraseña con límite de 3 intentos.
    Usa CACHE para persistencia (igual que con email de verificación)
    """
    from django.core.cache import cache
    import time
    
    email = (request.POST.get("email") or "").strip().lower()
    if not email:
        return JsonResponse({"ok": False, "error": "email_required"}, status=400)

    now = int(time.time())

    # ⭐ CONSTANTES
    MAX_RESENDS = 3
    LOCK_SECONDS = 60 * 30  # 30 minutos
    
    # ⭐ CLAVES CACHE
    count_key = f"pwreset:count:{email}"
    lock_key = f"pwreset:lock:{email}"

    # 1) ¿Está bloqueado?
    locked_until = cache.get(lock_key)
    if locked_until:
        remaining = int(locked_until - time.time())
        if remaining > 0:
            return JsonResponse({
                "ok": False,
                "locked": True,
                "retry_after": remaining,
                "error": f"Demasiados intentos. Intenta en {remaining // 60}:{remaining % 60:02d}."
            }, status=429)
        else:
            # Ya pasó el bloqueo, limpiar
            cache.delete(lock_key)
            cache.delete(count_key)

    # 2) ¿Cuántos reenvíos lleva?
    count = cache.get(count_key, 0)
    if count >= MAX_RESENDS:
        # Bloquear por 30 minutos
        locked_until_time = time.time() + LOCK_SECONDS
        cache.set(lock_key, locked_until_time, timeout=LOCK_SECONDS)
        return JsonResponse({
            "ok": False,
            "locked": True,
            "retry_after": LOCK_SECONDS,
            "error": "Demasiados intentos. Intenta en 30 minutos."
        }, status=429)

    # 3) Reenviar correo (Django password reset estándar)
    form = PasswordResetForm({"email": email})
    if form.is_valid():
        form.save(request=request)
    else:
        return JsonResponse({
            "ok": False,
            "error": "Email no válido o no registrado."
        }, status=400)

    # 4) Incrementar contador en CACHE (persiste 30 minutos)
    new_count = count + 1
    cache.set(count_key, new_count, timeout=LOCK_SECONDS)

    remaining = MAX_RESENDS - new_count
    
    return JsonResponse({
        "ok": True,
        "sent": True,
        "remaining": remaining,
        "message": f"Correo reenviado. Te quedan {remaining} intentos."
    }, status=200)

logger = logging.getLogger(__name__)

class PasswordChangeAuditView(PasswordChangeView):
    success_url = reverse_lazy("taxis:password_change_done")

    def form_valid(self, form):
        # Guarda el nuevo password y mantiene la sesión actual (comportamiento Django)
        response = super().form_valid(form)  # form.save() + update_session_auth_hash [web:290]

        # Auditoría real
        log_password_change(
            self.request,
            usuario_objetivo=self.request.user,
            accion="cambiar_password",
            descripcion="El usuario cambió su contraseña (sesión iniciada).",
        )

        # Logging opcional (no afecta BD)
        logger.info("AUDIT cambiar_password user_id=%s", self.request.user.id)
        return response


class PasswordResetConfirmAuditView(PasswordResetConfirmView):
    success_url = reverse_lazy("taxis:password_reset_complete")

    def form_valid(self, form):
        # Django aquí hace el reset y retorna el user (form.save()) y luego redirige [web:322]
        response = super().form_valid(form)  # guarda password / limpia session token [web:322]

        # Intentar obtener usuario objetivo de forma robusta
        usuario = getattr(self, "user", None)
        if usuario is None:
            # Fallback: si por alguna razón self.user no está, usa form.user si existe
            usuario = getattr(form, "user", None)

        if usuario is not None:
            log_password_change(
                self.request,
                usuario_objetivo=usuario,
                accion="reset_password",
                descripcion="Contraseña restablecida desde enlace de recuperación.",
            )
            logger.info("AUDIT reset_password user_id=%s", usuario.id)

        return response
    

# ====================================================================
# PERFIL PRESIDENTE
# ====================================================================
@login_required
def mi_perfil_presidente(request):
    if getattr(request.user, "role", "").upper() != "PRESIDENTE":
        return redirect("taxis:panel_general")

    docs_obj, _ = PresidenteIdentificacionDocs.objects.get_or_create(usuario=request.user)

    if request.method == "POST":
        action = request.POST.get("action")

        # ===== Guardar datos de perfil =====
        if action == "perfil":
            form = PresidentePerfilForm(request.POST, instance=request.user)
            docs_form = PresidenteIdentificacionDocsForm(instance=docs_obj)

            if form.is_valid():
                form.save()
                registrar_movimiento(request, accion="Actualizó su perfil", modulo="Perfiles")  # ya lo haces [file:16]
                messages.success(request, "Perfil actualizado correctamente.")
                return redirect("taxis:mi_perfil_presidente")

        # ===== Guardar documentos (dropzones) =====
        elif action == "docs":
            form = PresidentePerfilForm(instance=request.user)
            before_empty = (not docs_obj.cedula_frente) and (not docs_obj.cedula_detras) and (not docs_obj.rif)

            docs_form = PresidenteIdentificacionDocsForm(request.POST, request.FILES, instance=docs_obj)

            if docs_form.is_valid():
                obj = docs_form.save(commit=False)

                # Mantener anterior si no llegó archivo nuevo (igual que cooperativa docs) [file:16]
                if "cedula_frente" not in request.FILES:
                    obj.cedula_frente = docs_obj.cedula_frente
                if "cedula_detras" not in request.FILES:
                    obj.cedula_detras = docs_obj.cedula_detras
                if "rif" not in request.FILES:
                    obj.rif = docs_obj.rif

                obj.updated_by = request.user
                obj.save()

                uploaded = []
                if "cedula_frente" in request.FILES: uploaded.append("Cédula (frente)")
                if "cedula_detras" in request.FILES: uploaded.append("Cédula (detrás)")
                if "rif" in request.FILES: uploaded.append("RIF")

                if uploaded:
                    accion = "configurar"  # <- corto, dentro de ACCION_CHOICES
                    descripcion = "Actualizó documentos de identificación. Subidos: " + ", ".join(uploaded)

                    registrar_movimiento(
                     request,
                     accion=accion,
                     modulo="Perfiles",
                     descripcion=descripcion,
                     usuario=request.user,
                    )


                messages.success(request, "Documentos actualizados correctamente.")
                return redirect("taxis:mi_perfil_presidente")

        else:
            form = PresidentePerfilForm(instance=request.user)
            docs_form = PresidenteIdentificacionDocsForm(instance=docs_obj)
            messages.warning(request, "Acción no válida.")

    else:
        form = PresidentePerfilForm(instance=request.user)
        docs_form = PresidenteIdentificacionDocsForm(instance=docs_obj)

    movimientos = MovimientoAudit.objects.filter(usuario=request.user).order_by("-id")[:8]

    return render(request, "taxis/mi_perfil_presidente.html", {
        "form": form,
        "docs_form": docs_form,
        "docs_obj": docs_obj,
        "movimientos": movimientos,
    })


@login_required
@require_POST
def cooperativa_docs_update(request):
    if getattr(request.user, "role", "").upper() != "PRESIDENTE":
        messages.error(request, "No tienes permisos para actualizar documentos.")
        return redirect("taxis:panel_general")

    docs = CooperativaLegalDocs.get_solo()

    # --- estado ANTES (para saber si es primera vez o edición) ---
    before_acta = bool(getattr(docs, "acta_constitutiva_estatutos", None))
    before_asam = bool(getattr(docs, "acta_asamblea_extraordinaria", None))
    was_empty = (not before_acta) and (not before_asam)

    form = CooperativaLegalDocsForm(request.POST, request.FILES, instance=docs)

    # permitir submit sin re-subir archivos
    for f in ("acta_constitutiva_estatutos", "acta_asamblea_extraordinaria"):
        if f in form.fields:
            form.fields[f].required = False

    if form.is_valid():
        obj = form.save(commit=False)

        # mantener anterior si no llegó uno nuevo
        if "acta_constitutiva_estatutos" not in request.FILES:
            obj.acta_constitutiva_estatutos = docs.acta_constitutiva_estatutos

        if "acta_asamblea_extraordinaria" not in request.FILES:
            obj.acta_asamblea_extraordinaria = docs.acta_asamblea_extraordinaria

        if hasattr(obj, "updated_by"):
            obj.updated_by = request.user

        obj.save()

        # --- AUDITORÍA: solo si hubo upload real ---
        uploaded = []
        if "acta_constitutiva_estatutos" in request.FILES:
            uploaded.append("Acta Constitutiva y Estatutos")
        if "acta_asamblea_extraordinaria" in request.FILES:
            uploaded.append("Acta de Asamblea Extraordinaria")

        if uploaded:
            accion = "Agregó documentos legales de la cooperativa" if was_empty else "Actualizó documentos legales de la cooperativa"
            # Lo enganchas a Perfiles para no crear módulo nuevo
            registrar_movimiento(
                request,
                accion=accion,
                modulo="perfiles",
                descripcion="; ".join(uploaded),
                usuario=request.user,
            )

        if not request.FILES:
            messages.info(request, "No seleccionaste archivos nuevos; se mantuvieron los documentos actuales.")
        else:
            messages.success(request, "Documentos legales actualizados correctamente.")

        return redirect("taxis:panel_general")

    # errores -> edición
    for field, errors in form.errors.items():
        for error in errors:
            if field == "__all__":
                messages.error(request, str(error))
            else:
                label = form.fields.get(field).label if field in form.fields else field
                messages.error(request, f"{label}: {error}")

    return redirect(reverse("taxis:panel_general") + "?edit=1")